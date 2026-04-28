"""
Service d'import d'entreprises depuis Excel (XLSX).

Objectif:
- lire un fichier Excel fourni par l'API,
- normaliser les colonnes attendues (ex: phone_nu -> phone_number),
- n'importer que les lignes SANS website,
- dédupliquer (dans le fichier + en base),
- tracer le résultat dans des tables techniques d'import,
- déclencher l'analyse OSINT des numéros via le pipeline existant.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

from loguru import logger
from openpyxl import load_workbook
from sqlalchemy import delete as sa_delete
from sqlalchemy.orm import Session

from backend.database.models import (
    Entreprise,
    EntrepriseImportBatch,
    EntrepriseImportRow,
    EntreprisePhoneAnalysis,
    EntrepriseCategory,
    entreprise_category_links,
)
from backend.osint.services import PhoneOsintService
from backend.services.french_phone_detector import FrenchPhoneDetector


_SPACE_RE = re.compile(r"\s+")
_DIGITS_RE = re.compile(r"\D+")
_CATEGORY_SPLIT_RE = re.compile(r"[;,|/]+")
_CITY_FROM_ADDRESS_RE = re.compile(r"\b\d{5}\s+([A-Za-zÀ-ÖØ-öø-ÿ' -]{2,})\b")

# Catégories "bruit" à ignorer (sources type Google places)
_USELESS_CATEGORIES = {
    "etablissement",
    "établissement",
    "etablissement public",
    "établissement public",
    "point d'interet",
    "point d'intérêt",
    "point dinteret",
    "point d interet",
    "point d interêt",
    "point dinterêt",
    "poi",
}


def _as_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    s_lower = s.lower()
    if s_lower in ("nan", "none", "null", "n/a"):
        return None
    return s


def _normalize_header(header: Any) -> str:
    return (_as_str(header) or "").strip().lower()


def _phone_digits(phone: Optional[str]) -> Optional[str]:
    if not phone:
        return None
    digits = _DIGITS_RE.sub("", phone)
    # Normalisation FR: +33XXXXXXXXX -> 0XXXXXXXXX pour dedup cohérent
    if digits.startswith("33") and len(digits) == 11:
        digits = "0" + digits[2:]
    return digits or None


def _slugify(text: str) -> str:
    s = text.strip().lower()
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"[^a-z0-9\-]", "", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s or "category"


def _split_categories(raw: Optional[str]) -> List[str]:
    if not raw:
        return []
    parts = [_as_str(p) for p in _CATEGORY_SPLIT_RE.split(raw)]
    cleaned = [p for p in parts if p]
    seen = set()
    unique: List[str] = []
    for c in cleaned:
        key = c.lower()
        if key in _USELESS_CATEGORIES:
            continue
        if key in seen:
            continue
        seen.add(key)
        unique.append(c)
    return unique


def _is_empty_website(raw: Any) -> bool:
    s = _as_str(raw)
    return s is None


def _float_or_none(raw: Any) -> Optional[float]:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        try:
            return float(raw)
        except Exception:
            return None
    s = _as_str(raw)
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def _int_or_none(raw: Any) -> Optional[int]:
    if raw is None:
        return None
    if isinstance(raw, int):
        return raw
    if isinstance(raw, float):
        try:
            return int(raw)
        except Exception:
            return None
    s = _as_str(raw)
    if not s:
        return None
    try:
        return int(float(s.replace(",", ".")))
    except Exception:
        return None


def _extract_city_from_address(*parts: Optional[str]) -> Optional[str]:
    for part in parts:
        if not part:
            continue
        m = _CITY_FROM_ADDRESS_RE.search(part)
        if m:
            city = _as_str(m.group(1))
            if city:
                return city
    return None


@dataclass(frozen=True)
class ImportStats:
    batch_id: int
    original_filename: Optional[str]
    total_rows: int
    imported_rows: int
    skipped_with_website: int
    skipped_invalid: int
    skipped_duplicates: int


class EntrepriseImportService:
    """
    Service d'import XLSX.

    Le format exact des colonnes peut varier, on accepte des alias.
    """

    # Canonical -> aliases (headers lowercase)
    _COLUMN_ALIASES: Dict[str, List[str]] = {
        "name": ["name", "nom", "entreprise", "company", "raison sociale"],
        "website": ["website", "site", "site web", "url", "lien", "web"],
        "category": ["category", "categorie", "catégorie", "type"],
        "category_translate": ["category_translate", "category translate", "categorie_traduite", "categorie translate"],
        "phone_number": ["phone_number", "phone", "telephone", "téléphone", "phone_nu", "phone nu", "phone_nu."],
        "country": ["country", "pays"],
        "city": ["city", "ville", "commune", "locality"],
        "address_1": ["address_1", "adresse_1", "adresse 1", "address", "adresse", "address_full", "adresse complète"],
        "address_2": ["address_2", "adresse_2", "adresse 2", "complément", "complement"],
        "longitude": ["longitude", "lng", "long"],
        "latitude": ["latitude", "lat"],
        "rating": ["rating", "note", "note_google", "google rating"],
        "reviews_count": ["reviews_count", "reviews", "avis", "nb_avis", "nb_avis_google", "review count"],
    }

    def __init__(self, db: Session) -> None:
        self._db = db
        self._french_detector: Optional[FrenchPhoneDetector] = None

    def import_xlsx(
        self,
        xlsx_bytes: bytes,
        original_filename: Optional[str] = None,
        source: str = "excel",
        analyze_phone: bool = True,
        batch_id: Optional[int] = None,
        progress_callback: Optional[Callable[[Dict[str, Any]], None]] = None,
    ) -> ImportStats:
        """
        Importe un XLSX et renvoie un résumé.
        """
        if batch_id is not None:
            batch = self._db.get(EntrepriseImportBatch, batch_id)  # type: ignore[arg-type]
            if batch is None:
                raise ValueError(f"Batch introuvable: {batch_id}")
            if original_filename:
                batch.original_filename = original_filename
            batch.source = source
            self._db.commit()
        else:
            batch = EntrepriseImportBatch(
                original_filename=original_filename,
                source=source,
                total_rows=0,
                imported_rows=0,
                skipped_with_website=0,
                skipped_invalid=0,
                skipped_duplicates=0,
            )
            self._db.add(batch)
            self._db.commit()
            self._db.refresh(batch)

        wb = load_workbook(filename=BytesIO(xlsx_bytes), read_only=True, data_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            logger.warning("Import XLSX: fichier vide")
            return ImportStats(
                batch_id=batch.id,
                original_filename=original_filename,
                total_rows=0,
                imported_rows=0,
                skipped_with_website=0,
                skipped_invalid=0,
                skipped_duplicates=0,
            )

        header_to_index: Dict[str, int] = {}
        for idx, h in enumerate(header_row):
            key = _normalize_header(h)
            if key:
                header_to_index[key] = idx

        col_index = self._resolve_columns(header_to_index)
        logger.info(f"Import XLSX: colonnes resolues={col_index}")

        # On connaît le nombre de lignes "data" à l'avance (pour jauge), sans charger toute la feuille.
        # openpyxl read_only ne supporte pas len(rows_iter), mais ws.max_row donne une borne utile.
        try:
            # -1 pour l'en-tête
            batch.total_rows = max(0, int(getattr(ws, "max_row", 1)) - 1)
            self._db.commit()
        except Exception:
            pass

        # Dedup in-file
        seen_signatures: set[Tuple[str, str]] = set()

        total_rows = 0
        for excel_row_number, row_values in enumerate(rows_iter, start=2):
            total_rows += 1
            normalized = self._row_to_record(row_values, col_index)

            row_trace = EntrepriseImportRow(
                batch_id=batch.id,
                row_number=excel_row_number,
                name=normalized.get("name"),
                website=normalized.get("website"),
                phone_number=normalized.get("phone_number"),
                country=normalized.get("country"),
                address_1=normalized.get("address_1"),
                address_2=normalized.get("address_2"),
                category=normalized.get("category"),
                status="pending",
            )
            self._db.add(row_trace)

            if progress_callback and (total_rows == 1 or total_rows % 25 == 0):
                # Event "heartbeat" périodique pour la jauge, même si beaucoup de lignes sont ignorées.
                try:
                    progress_callback(
                        {
                            "batch_id": batch.id,
                            "current": total_rows,
                            "imported_rows": batch.imported_rows,
                            "skipped_with_website": batch.skipped_with_website,
                            "skipped_invalid": batch.skipped_invalid,
                            "skipped_duplicates": batch.skipped_duplicates,
                        }
                    )
                except Exception:
                    pass

            name = normalized.get("name")
            website = normalized.get("website")
            phone = normalized.get("phone_number")

            if not name:
                row_trace.status = "skipped_invalid"
                row_trace.reason = "Nom manquant"
                batch.skipped_invalid += 1
                continue

            # Regle metier: ne garder que sans website
            if website:
                row_trace.status = "skipped_website"
                row_trace.reason = "Website present"
                batch.skipped_with_website += 1
                continue

            # Signature dedup: priorite phone_digits, sinon name+adresse
            digits = _phone_digits(phone) or ""
            addr_sig = _SPACE_RE.sub(" ", (normalized.get("address_1") or "").strip().lower())
            sig: Tuple[str, str]
            if digits:
                sig = ("phone", digits)
            else:
                sig = ("name_addr", f"{name.strip().lower()}|{addr_sig}")

            if sig in seen_signatures and sig[1]:
                row_trace.status = "skipped_duplicate"
                row_trace.reason = "Doublon dans le fichier"
                batch.skipped_duplicates += 1
                continue
            seen_signatures.add(sig)

            # Dedup DB: simple et tolerant
            existing = self._find_existing_entreprise(phone_digits=digits or None, name=name)
            if existing is not None:
                row_trace.status = "skipped_duplicate"
                row_trace.reason = f"Deja en base (entreprise_id={existing.id})"
                row_trace.entreprise_id = existing.id
                batch.skipped_duplicates += 1
                continue

            entreprise = Entreprise(
                name=name.strip(),
                website=website,
                phone_number=phone,
                phone_digits=digits or None,
                country=normalized.get("country"),
                city=normalized.get("city"),
                address_1=normalized.get("address_1"),
                address_2=normalized.get("address_2"),
                longitude=_float_or_none(normalized.get("longitude")),
                latitude=_float_or_none(normalized.get("latitude")),
                rating=_float_or_none(normalized.get("rating")),
                reviews_count=_int_or_none(normalized.get("reviews_count")),
            )
            # Ajouter l'entreprise en session avant les liens M2M pour éviter les warnings SQLAlchemy.
            self._db.add(entreprise)
            # Force l'ID maintenant puis purge défensive des anciens liens orphelins potentiels.
            self._db.flush()
            self._db.execute(
                sa_delete(entreprise_category_links).where(
                    entreprise_category_links.c.entreprise_id == entreprise.id
                )
            )

            # Categorie_translate peut contenir plusieurs categories -> normalisation M2M
            linked_category_ids: set[int] = set()
            for category_name in _split_categories(normalized.get("category_translate")):
                category = self._get_or_create_category(category_name)
                # Evite les doublons M2M quand plusieurs libellés mappent vers le même slug/catégorie.
                if category.id in linked_category_ids:
                    continue
                linked_category_ids.add(category.id)
                entreprise.categories.add(category)

            self._db.commit()
            self._db.refresh(entreprise)

            row_trace.status = "imported"
            row_trace.reason = None
            row_trace.entreprise_id = entreprise.id
            batch.imported_rows += 1

            if analyze_phone and phone:
                self._enqueue_phone_analysis(entreprise_id=entreprise.id, phone_number=phone)

            if progress_callback:
                try:
                    progress_callback(
                        {
                            "batch_id": batch.id,
                            "current": total_rows,
                            "imported_rows": batch.imported_rows,
                            "skipped_with_website": batch.skipped_with_website,
                            "skipped_invalid": batch.skipped_invalid,
                            "skipped_duplicates": batch.skipped_duplicates,
                        }
                    )
                except Exception:
                    # Ne jamais casser l'import si le callback échoue
                    pass

        batch.total_rows = total_rows
        self._db.commit()

        return ImportStats(
            batch_id=batch.id,
            original_filename=original_filename,
            total_rows=batch.total_rows,
            imported_rows=batch.imported_rows,
            skipped_with_website=batch.skipped_with_website,
            skipped_invalid=batch.skipped_invalid,
            skipped_duplicates=batch.skipped_duplicates,
        )

    def _resolve_columns(self, header_to_index: Dict[str, int]) -> Dict[str, Optional[int]]:
        resolved: Dict[str, Optional[int]] = {}
        for canonical, aliases in self._COLUMN_ALIASES.items():
            idx = None
            for a in aliases:
                a_norm = a.strip().lower()
                if a_norm in header_to_index:
                    idx = header_to_index[a_norm]
                    break
            resolved[canonical] = idx
        return resolved

    def _row_to_record(self, row_values: Tuple[Any, ...], col_index: Dict[str, Optional[int]]) -> Dict[str, Any]:
        def get_col(col: str) -> Any:
            idx = col_index.get(col)
            if idx is None:
                return None
            if idx < 0 or idx >= len(row_values):
                return None
            return row_values[idx]

        name = _as_str(get_col("name"))
        website_raw = get_col("website")
        website = _as_str(website_raw) if not _is_empty_website(website_raw) else None
        phone = _as_str(get_col("phone_number"))
        category = _as_str(get_col("category"))
        category_translate = _as_str(get_col("category_translate"))
        country = _as_str(get_col("country"))
        city = _as_str(get_col("city"))
        address_1 = _as_str(get_col("address_1"))
        address_2 = _as_str(get_col("address_2"))
        if not city:
            city = _extract_city_from_address(address_2, address_1)
        if phone:
            detected_city = self._extract_city_from_phone(phone)
            if detected_city:
                # Priorise la reference telecom (french_phone_data) si ville absente
                # ou si la ville extraite de l'adresse est potentiellement bruitée.
                if not city or len(city.strip()) < 3:
                    city = detected_city

        return {
            "name": name,
            "website": website,
            "phone_number": phone,
            "category": category,
            "category_translate": category_translate,
            "country": country,
            "city": city,
            "address_1": address_1,
            "address_2": address_2,
            "longitude": get_col("longitude"),
            "latitude": get_col("latitude"),
            "rating": get_col("rating"),
            "reviews_count": get_col("reviews_count"),
        }

    def _extract_city_from_phone(self, phone_number: str) -> Optional[str]:
        try:
            if self._french_detector is None:
                self._french_detector = FrenchPhoneDetector(db=self._db)
            result = self._french_detector.detect(phone_number)
            city = _as_str(result.get("city"))
            return city
        except Exception:
            return None

    def _find_existing_entreprise(self, phone_digits: Optional[str], name: str) -> Optional[Entreprise]:
        q = self._db.query(Entreprise)
        if phone_digits:
            existing = q.filter(Entreprise.phone_digits == phone_digits).one_or_none()
            if existing:
                return existing
        # Fallback tolerant sur nom (si pas de tel)
        name_norm = name.strip().lower()
        return (
            q.filter(Entreprise.name.ilike(name_norm))  # type: ignore[attr-defined]
            .order_by(Entreprise.created_at.desc())
            .first()
        )

    def _enqueue_phone_analysis(self, entreprise_id: int, phone_number: str) -> None:
        digits = _phone_digits(phone_number)
        analysis = EntreprisePhoneAnalysis(
            entreprise_id=entreprise_id,
            phone_number=phone_number,
            phone_digits=digits,
            status="queued",
        )
        self._db.add(analysis)
        self._db.commit()
        self._db.refresh(analysis)

        try:
            phone_osint = PhoneOsintService(self._db)
            profile = phone_osint.ensure_profile_for_number(phone_number=phone_number, caller_id=None)
            analysis.phone_profile_id = profile.id
            analysis.status = "queued"
            self._db.commit()
        except Exception as exc:
            logger.warning(f"Analyse OSINT impossible pour entreprise_id={entreprise_id}: {exc}")
            analysis.status = "failed"
            analysis.error_message = str(exc)[:500]
            self._db.commit()

    def _get_or_create_category(self, category_name: str) -> EntrepriseCategory:
        normalized = category_name.strip()
        slug = _slugify(normalized)
        # Evite les autoflush intermédiaires pendant la construction de l'entreprise + ses liens M2M.
        with self._db.no_autoflush:
            existing = (
                self._db.query(EntrepriseCategory)
                .filter(EntrepriseCategory.slug == slug)
                .one_or_none()
            )
        if existing:
            return existing
        category = EntrepriseCategory(name=normalized, slug=slug)
        self._db.add(category)
        # Important: ne pas commit ici pendant l'import M2M, sinon flush partiel
        # et risque de doublons d'insert dans la table de liaison.
        self._db.flush()
        return category

